"""CSV → XLSX: боевой файл для обзвона B2B-базы ритуальных компаний.

Структура книги:
  «Сводка»           — дашборд: объём, заполненность контактов, источники, города, типы
  «Все ритуальные»   — все записи, кроме флористики (главный лист для обзвона)
  «Флористика»       — client_type = флористика
  «Требуют проверки» — comment содержит needs_review (ambiguous-классификация)
  <Город>            — отдельный лист на каждый город с ≥5 записями
  «Остальные»        — города с <5 записями вместе
  «Без контактов»    — записи без phone и без email (для ручного добивания)

Каждый лист обзвона:
  - № (порядковый), затем Город/Название/Тип/Телефон/Email/Адрес/Сайт/Соцсеть/…
  - «Статус обзвона» — пустая колонка с выпадающим списком (Да/Нет/Перезвонить/…)
  - сортировка: сначала записи с телефоном (это инструмент для звонков),
    затем с email, внутри — по городу и названию
  - баннер-заголовок с числом записей, заморозка шапки, автофильтр
  - заливка строки по client_type, кликабельные phone(tel:)/email(mailto:)/site
"""
from __future__ import annotations

import csv
import os
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Колонки листа обзвона: (заголовок, csv-поле или None для генерируемых, ширина).
# Порядок «для звонящего»: город → название → тип → ТЕЛЕФОН → email → остальное.
COLUMNS: list[tuple[str, str | None, int]] = [
    ("№",              None,          5),
    ("Город",          "city",        15),
    ("Название",       "name",        38),
    ("Тип клиента",    "client_type", 20),
    ("Телефон",        "phone",       20),
    ("Email",          "email",       26),
    ("Адрес",          "address",     38),
    ("Сайт",           "website",     30),
    ("Соцсеть",        "social",      24),
    ("Категория",      "category",    18),
    ("Источник",       "source",      12),
    ("Статус обзвона", None,          16),
    ("Комментарий",    "comment",     26),
]

HEADERS = [c[0] for c in COLUMNS]
N_COLS = len(COLUMNS)
LAST_COL = get_column_letter(N_COLS)

# Заливка строки по client_type (мягкие пастельные тона)
FILL_BY_TYPE = {
    "ритуальное агентство":  "DCEEFB",  # голубой
    "похоронное бюро":       "E2D9F3",  # сиреневый
    "ритуальный магазин":    "DCFCE7",  # зелёный
    "мастерская памятников": "FEF3C7",  # жёлтый
    "кладбищенские услуги":  "FFE4E6",  # розовый
    "флористика":            "FFF7ED",  # персиковый
    "прочее":                "FFFFFF",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")   # тёмно-серый
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BANNER_FILL = PatternFill("solid", fgColor="0F766E")   # бирюзовый баннер
BANNER_FONT = Font(bold=True, color="FFFFFF", size=13)
LINK_FONT = Font(color="0563C1", underline="single")
NO_CONTACT_FILL = PatternFill("solid", fgColor="FECACA")  # светло-красный

SOURCE_FONT_COLOR = {
    "OSM": "1E40AF", "VK": "2563EB", "Я.Карты": "B91C1C", "Поиск": "047857",
    "Wikipedia": "9333EA", "Wikidata": "374151", "Crawler": "C2410C",
}

WRAP_HEADERS = {"Адрес", "Комментарий"}
_THIN = Side(style="thin", color="D1D5DB")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Варианты статуса обзвона — выпадающий список в колонке «Статус обзвона»
CALL_STATUSES = "Да,Нет,Перезвонить,Не дозвонился,Отказ,Не актуально"


def _read_csv(path: str) -> list[dict]:
    rows: list[dict] = []
    for delim in (";", ","):
        try:
            with open(path, encoding="utf-8-sig") as f:
                sample = f.read(2048)
                f.seek(0)
                if delim in sample:
                    reader = csv.DictReader(f, delimiter=delim)
                    rows = list(reader)
                    if rows and "name" in rows[0]:
                        return rows
        except Exception:
            continue
    return rows


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel запрещает : \\ / ? * [ ] в именах листов и >31 символа; имена уникальны."""
    bad = set(":\\/?*[]")
    base = "".join(" " if c in bad else c for c in name).strip()[:31] or "Лист"
    out = base
    i = 2
    while out in used:
        suffix = f" {i}"
        out = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(out)
    return out


def _sort_rows(rows: list[dict]) -> list[dict]:
    """Инструмент для обзвона → сначала с телефоном, потом с email, по городу+названию."""
    def key(r: dict):
        has_phone = 0 if (r.get("phone") or "").strip() else 1
        has_email = 0 if (r.get("email") or "").strip() else 1
        city = (r.get("city") or "").strip().lower()
        name = (r.get("name") or "").strip().lower()
        return (has_phone, has_email, city, name)
    return sorted(rows, key=key)


def _write_sheet(ws, title: str, rows: list[dict], with_filter: bool = True) -> None:
    """Лист обзвона: баннер + шапка + данные + статус-дропдаун + оформление."""
    rows = _sort_rows(rows)

    # строка 1 — баннер с названием листа и числом записей
    ws.merge_cells(f"A1:{LAST_COL}1")
    banner = ws.cell(row=1, column=1, value=f"{title}   ·   записей: {len(rows)}")
    banner.fill = BANNER_FILL
    banner.font = BANNER_FONT
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # строка 2 — шапка
    for c_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=c_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    # данные с 3-й строки
    for i, row in enumerate(rows):
        r_idx = i + 3
        client_type = (row.get("client_type") or "прочее").lower()
        no_phone = not (row.get("phone") or "").strip()
        no_email = not (row.get("email") or "").strip()
        if no_phone and no_email:
            row_fill = NO_CONTACT_FILL
        else:
            color = FILL_BY_TYPE.get(client_type, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=color) if color != "FFFFFF" else None

        for c_idx, (header, field, _w) in enumerate(COLUMNS, start=1):
            if header == "№":
                val = i + 1
            elif field is None:
                val = ""  # «Статус обзвона» — пустая для ручного заполнения
            else:
                val = row.get(field, "") or ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header in WRAP_HEADERS,
                horizontal="center" if header == "№" else "left",
            )
            if row_fill is not None:
                cell.fill = row_fill

            sval = str(val)
            if header == "Сайт" and sval.startswith(("http://", "https://")):
                cell.hyperlink = sval
                cell.font = LINK_FONT
            elif header == "Email" and "@" in sval:
                cell.hyperlink = f"mailto:{sval}"
                cell.font = LINK_FONT
            elif header == "Телефон" and sval:
                tel = "".join(ch for ch in sval if ch.isdigit() or ch == "+")
                if tel:
                    cell.hyperlink = f"tel:{tel}"
                    cell.font = LINK_FONT
            elif header == "Соцсеть" and sval.startswith(("http://", "https://")):
                cell.hyperlink = sval
                cell.font = LINK_FONT
            elif header == "Источник" and sval:
                cell.font = Font(bold=True, color=SOURCE_FONT_COLOR.get(sval, "374151"))

    # ширина колонок
    for c_idx, (_h, _f, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # заморозка: шапка (строки 1-2) + первые 3 колонки (№/Город/Название)
    ws.freeze_panes = "D3"

    if rows and with_filter:
        ws.auto_filter.ref = f"A2:{LAST_COL}{len(rows) + 2}"

    # выпадающий список в «Статус обзвона»
    if rows:
        status_col = get_column_letter(HEADERS.index("Статус обзвона") + 1)
        dv = DataValidation(type="list", formula1=f'"{CALL_STATUSES}"', allow_blank=True)
        dv.add(f"{status_col}3:{status_col}{len(rows) + 2}")
        ws.add_data_validation(dv)


def _write_summary(ws, rows: list[dict], src_csv: str) -> None:
    """Дашборд: ключевые цифры базы, заполненность контактов, источники, города, типы."""
    total = len(rows)

    def cnt(field: str) -> int:
        return sum(1 for r in rows if (r.get(field) or "").strip())

    with_phone, with_email = cnt("phone"), cnt("email")
    with_addr, with_site, with_social = cnt("address"), cnt("website"), cnt("social")
    callable_now = sum(1 for r in rows
                       if (r.get("phone") or "").strip() or (r.get("email") or "").strip())

    src_cnt = Counter((r.get("source") or "—") for r in rows)
    city_cnt = Counter((r.get("city") or "—").strip() or "—" for r in rows)
    type_cnt = Counter((r.get("client_type") or "—") for r in rows)

    def pct(n: int) -> str:
        return f"{100 * n / total:.0f}%" if total else "—"

    title_font = Font(bold=True, size=14, color="0F766E")
    sec_font = Font(bold=True, size=11, color="1F2937")
    sec_fill = PatternFill("solid", fgColor="E5E7EB")
    bold = Font(bold=True)

    ws.cell(row=1, column=1, value="РИТУАЛЬНЫЙ B2B — сводка по базе").font = title_font
    ws.cell(row=2, column=1, value=f"Файл: {os.path.basename(src_csv)}")
    ws.cell(row=3, column=1, value=f"Сформирован: {datetime.now():%Y-%m-%d %H:%M}")

    r = 5

    def section(name: str) -> None:
        nonlocal r
        c = ws.cell(row=r, column=1, value=name)
        c.font = sec_font
        c.fill = sec_fill
        ws.cell(row=r, column=2).fill = sec_fill
        r += 1

    def line(k: str, v: str, indent: bool = False) -> None:
        nonlocal r
        kc = ws.cell(row=r, column=1, value=("   " + k) if indent else k)
        if not indent:
            kc.font = bold
        ws.cell(row=r, column=2, value=v)
        r += 1

    section("Объём базы")
    line("Всего записей", str(total))
    line("Готовы к обзвону (есть телефон или email)", f"{callable_now} ({pct(callable_now)})")
    r += 1

    section("Заполненность контактов")
    line("С телефоном", f"{with_phone} ({pct(with_phone)})", indent=True)
    line("С email", f"{with_email} ({pct(with_email)})", indent=True)
    line("С адресом", f"{with_addr} ({pct(with_addr)})", indent=True)
    line("С сайтом", f"{with_site} ({pct(with_site)})", indent=True)
    line("С соцсетью", f"{with_social} ({pct(with_social)})", indent=True)
    r += 1

    section("Типы клиентов")
    for k, v in type_cnt.most_common():
        line(k, str(v), indent=True)
    r += 1

    section("Источники")
    for k, v in src_cnt.most_common():
        line(k, str(v), indent=True)
    r += 1

    section("Города (топ 25)")
    for k, v in city_cnt.most_common(25):
        line(k, str(v), indent=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.sheet_view.showGridLines = False


def build_xlsx(csv_path: str, xlsx_path: str | None = None) -> str | None:
    """Построить XLSX из CSV. Возвращает путь к файлу."""
    if not os.path.exists(csv_path):
        print(f"[xlsx] CSV не найден: {csv_path}")
        return None

    rows = _read_csv(csv_path)
    if not rows:
        print(f"[xlsx] CSV пуст: {csv_path}")
        return None

    if not xlsx_path:
        xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    _write_summary(ws_summary, rows, csv_path)

    used_names = {"Сводка"}

    def _is_review(r: dict) -> bool:
        return "review" in (r.get("comment") or "").lower()

    florist_rows = [r for r in rows if (r.get("client_type") or "").lower() == "флористика"]
    review_rows = [r for r in rows if _is_review(r)
                   and (r.get("client_type") or "").lower() != "флористика"]
    # Главный лист обзвона — подтверждённый ритуал: не флористика и НЕ ambiguous.
    # Ambiguous (needs_review) уходят в отдельный лист, чтобы не мусорить обзвон.
    ritual_rows = [r for r in rows
                   if (r.get("client_type") or "").lower() != "флористика"
                   and not _is_review(r)]

    name = _safe_sheet_name("Все ритуальные", used_names)
    _write_sheet(wb.create_sheet(name), "Все ритуальные (подтверждённые)", ritual_rows)

    if florist_rows:
        name = _safe_sheet_name("Флористика", used_names)
        _write_sheet(wb.create_sheet(name), "Флористика", florist_rows)

    if review_rows:
        name = _safe_sheet_name("Требуют проверки", used_names)
        _write_sheet(wb.create_sheet(name),
                     "Требуют ручной проверки (VK, неоднозначные)", review_rows)

    # По городам и «без контактов» — только подтверждённые (ambiguous живут
    # лишь в «Требуют проверки»), чтобы городские листы были чистым обзвоном.
    core_rows = [r for r in rows if not _is_review(r)]

    # По городам: отдельный лист на каждый с ≥5 записями, мелкие → «Остальные»
    MIN_CITY_ROWS = 5
    by_city: dict[str, list[dict]] = defaultdict(list)
    for r in core_rows:
        city = (r.get("city") or "Не указан").strip() or "Не указан"
        by_city[city].append(r)

    others: list[dict] = []
    for city in sorted(by_city, key=lambda c: -len(by_city[c])):
        if len(by_city[city]) < MIN_CITY_ROWS:
            others.extend(by_city[city])
            continue
        name = _safe_sheet_name(city, used_names)
        _write_sheet(wb.create_sheet(name), f"Город: {city}", by_city[city])

    if others:
        name = _safe_sheet_name("Остальные", used_names)
        _write_sheet(wb.create_sheet(name), "Остальные города (<5 записей)", others)

    no_contacts = [r for r in core_rows
                   if not (r.get("phone") or "").strip()
                   and not (r.get("email") or "").strip()]
    if no_contacts:
        name = _safe_sheet_name("Без контактов", used_names)
        _write_sheet(wb.create_sheet(name), "Без контактов (нужно добить вручную)", no_contacts)

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"[xlsx] сохранён: {xlsx_path} (листов: {len(wb.sheetnames)})")
    return xlsx_path
